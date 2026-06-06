"""
ingest.py — Altis Groep cash-flow data foundation
==================================================

Converts raw accounting exports (the .xlsx files in this folder) into the
ONE unified `cash_event` schema that every dashboard/role reads from.

This is deliverable #01: "Data ingestion from at least one accounting system
and reconciliation into a unified schema." The output array has exactly the
same shape as the hand-written CASH_EVENTS seed the UI team is building
against, so it is a drop-in replacement — nothing in the UI needs to change.

The schema (one row = one cash movement):
    event_id        unique id, also a stable audit handle
    week            1..13  (week the cash moves, within the forecast horizon)
    opco            operating company
    driver          materials | subcontractor | milestone_billing |
                    payment_lag | weather
    amount          signed euros. negative = outflow, positive = inflow
    source_system   gilde | yuki | exact | snelstart
    source_row_id   pointer back to the exact raw row  (audit trail)
    assumptions     tags explaining why this row looks the way it does
    scenario        base | wet | dry
  + audit extras (date, gl_account, gl_label, vat_rate, net_amount, debet,
    credit, journal, doc_no, counterparty, description, source_file,
    source_excel_row) — the UI ignores these, a controller lives in them.

Pipeline (mirrors the brief's required layering):
    ingest  ->  normalize  ->  reconcile  ->  window to 13 weeks  ->  serialise

Run:
    python ingest.py --in /path/to/uploads --out /path/to/outputs
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import glob
from collections import defaultdict

from openpyxl import load_workbook


# ----------------------------------------------------------------------------
# CONFIG — everything a controller would tune lives here, not buried in code.
# ----------------------------------------------------------------------------

# Chart of accounts -> driver / VAT / label.
# This IS the GL mapping. Adding a new account is a one-line edit here, not a
# code change (edge case: "new GL account" -> model adapts, doesn't break).
#
# All accounts in the supplied exports are REVENUE (omzet) accounts, so they
# map to milestone_billing. The materials / subcontractor rows below are
# wired-but-unused placeholders: the moment a cost-account export is dropped
# in, it routes to the right driver with zero code change.
GL_ACCOUNTS = {
    # account : (driver,              vat_rate, label)
    8000: ("milestone_billing", 0.21, "Omzet hoog (21%)"),
    8001: ("milestone_billing", 0.00, "Omzet verlegd (BTW verlegd / reverse charge)"),
    8002: ("milestone_billing", 0.09, "Omzet laag (9%)"),
    8004: ("milestone_billing", 0.00, "Omzet 0% / niet bij u belast"),
    8005: ("milestone_billing", 0.00, "Omzet heffing verlegd (reverse charge)"),
    # ---- placeholders, route correctly the day a cost export arrives ----
    # 7000: ("materials",     0.21, "Inkoop dakmateriaal"),
    # 4400: ("subcontractor", 0.21, "Onderaannemers"),
}

# Source system inferred from the export LAYOUT. The two layouts here come
# from different systems; these defaults are an inference from vocabulary
# ("Administratie/Boekjaar/Bkst.nr." == Exact) and should be confirmed with
# Altis. Override per-file via SOURCE_SYSTEM_BY_FILE if you know better.
SOURCE_SYSTEM_BY_FORMAT = {
    "fintransactions": "exact",     # Exact-style report export
    "gb": "snelstart",              # grootboek export (inferred — confirm)
}
SOURCE_SYSTEM_BY_FILE: dict[str, str] = {}   # e.g. {"GB_8001_jan-dec_23.xlsx": "yuki"}

# Short audit prefix per source system, used in source_row_id.
SYSTEM_PREFIX = {"exact": "EX", "snelstart": "SS", "yuki": "YK", "gilde": "GI"}

# Cash realism: a revenue GL line is booked EX-VAT, but the customer pays
# INCL-VAT. For a cash-flow forecast we gross up to what actually hits the
# bank. Reverse-charge / 0% accounts get factor 1.00 (correct Dutch
# treatment). Turn off with --no-gross-up to keep net (ex-VAT) figures.
APPLY_VAT_GROSS_UP = True

# Covenant + opening cash — placeholders, swap for the covenant-terms doc.
OPENING_CASH = 250000
COVENANT = {"type": "min_liquidity", "threshold": 150000, "warningBuffer": 50000}


# ----------------------------------------------------------------------------
# small helpers
# ----------------------------------------------------------------------------

def _rows(path: str) -> list[tuple]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def _cell(row, i):
    return row[i] if row and len(row) > i else None


def _num(v) -> float:
    return float(v) if isinstance(v, (int, float)) else 0.0


def _to_int_account(v):
    """'8000.0' / 8000 / '8000' -> 8000 ; anything weird -> None."""
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return None


def _to_date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    return None


def _doc_no(v) -> str:
    """Document number may arrive as int, float (2300379.0) or str (KA240004)."""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def classify(account: int):
    """account -> (driver, vat_rate, label). Unknown accounts do NOT crash:
    they fall through to an 'unmapped' driver and a loud assumption tag, so a
    controller sees them instead of the pipeline silently dropping money."""
    if account in GL_ACCOUNTS:
        return GL_ACCOUNTS[account]
    return ("unmapped", 0.0, f"ONGEMAPT grootboek {account} — controller: map in GL_ACCOUNTS")


# ----------------------------------------------------------------------------
# format detection
# ----------------------------------------------------------------------------

def detect_format(rows: list[tuple]) -> str:
    head = " ".join(str(_cell(rows[0], i) or "") for i in range(3)).lower() if rows else ""
    if head.startswith("rekening"):
        return "gb"
    if "administratie" in head:
        return "fintransactions"
    # fallback: a 'Nr.' data header => fintransactions
    for r in rows[:20]:
        if str(_cell(r, 0)).strip() == "Nr.":
            return "fintransactions"
    return "gb"


# ----------------------------------------------------------------------------
# parsers — each returns a list of normalised events (scenario-agnostic)
# ----------------------------------------------------------------------------

def _normalise(*, account, date, debet, credit, doc_no, journal, description,
               counterparty, source_system, source_file, excel_row, line_no,
               gross_up):
    """Build one unified event from a single raw ledger line."""
    driver, vat_rate, label = classify(account)

    net = round(_num(credit) - _num(debet), 2)        # +inflow / -outflow
    factor = (1 + vat_rate) if gross_up else 1.0
    amount = round(net * factor, 2)                    # what hits the bank

    prefix = SYSTEM_PREFIX.get(source_system, "XX")
    source_row_id = f"{prefix}-{doc_no}#L{line_no}"

    assumptions = [label, f"dagboek: {journal}" if journal else "dagboek: n/b"]
    if vat_rate and gross_up:
        assumptions.append(f"cash incl. {int(vat_rate*100)}% BTW (×{factor:g})")
    elif gross_up:
        assumptions.append("geen BTW op cashflow (verlegd / 0%)")
    text = (description or "").lower()
    if net < 0 or "corr" in text or "oninba" in text or "credit" in text:
        assumptions.append("correctie / creditnota (vermindert facturatie)")

    return {
        # ----- exact CASH_EVENTS schema (drop-in) -----
        "event_id": source_row_id,
        "week": None,                 # stamped later by window_to_horizon()
        "opco": "Dakdekkersbedrijf Peter Ummels",
        "driver": driver,
        "amount": amount,
        "source_system": source_system,
        "source_row_id": source_row_id,
        "assumptions": assumptions,
        "scenario": "base",
        # ----- audit extras (UI ignores; controller uses) -----
        "date": date.isoformat() if date else None,
        "gl_account": account,
        "gl_label": label,
        "vat_rate": vat_rate,
        "net_amount": net,
        "debet": round(_num(debet), 2),
        "credit": round(_num(credit), 2),
        "journal": journal,
        "doc_no": doc_no,
        "counterparty": counterparty,
        "description": description,
        "source_file": source_file,
        "source_excel_row": excel_row,
    }


def parse_fintransactions(rows, source_system, source_file, gross_up):
    """Exact-style export: one account for the whole file (in the header),
    data block under a 'Nr. | Per. | Datum | Bkst.nr. | Dagboek | Debet |
    Credit' header, terminated by Totaal / Eindsaldo."""
    # account from the criteria header (row 6, col 1): "8005 - omzet ..."
    account = None
    for r in rows[:12]:
        if str(_cell(r, 0)).strip().lower().startswith("grootboekrekening"):
            account = _to_int_account(str(_cell(r, 1)).split(" - ")[0])
            break
    # data header
    h = next((i for i, r in enumerate(rows) if str(_cell(r, 0)).strip() == "Nr."), None)
    if h is None:
        return [], None
    events, line = [], 0
    eindsaldo = None
    for i, r in enumerate(rows[h + 1:], start=h + 2):   # +2 -> 1-based excel row
        first = _cell(r, 0)
        s = "" if first is None else str(first).strip()
        if s in ("", "None", "Totaal"):
            continue
        if s == "Eindsaldo":
            eindsaldo = round(_num(_cell(r, 6)) - _num(_cell(r, 5)), 2) or _num(_cell(r, 6))
            continue
        line += 1
        events.append(_normalise(
            account=account,
            date=_to_date(_cell(r, 2)),
            debet=_cell(r, 5), credit=_cell(r, 6),
            doc_no=_doc_no(_cell(r, 3)),
            journal=str(_cell(r, 4)).strip() if _cell(r, 4) else "",
            description=str(_cell(r, 4)).strip() if _cell(r, 4) else "",
            counterparty=None,
            source_system=source_system, source_file=source_file,
            excel_row=i, line_no=line, gross_up=gross_up,
        ))
    return events, eindsaldo


def parse_gb(rows, source_system, source_file, gross_up):
    """Grootboek export: account per row. Columns:
    Rekening | Periode | Datum | Boeknummer | Trek | Debet | Credit |
    Boekingstekst | Dagboek | BTW | BTW-srt."""
    events, line = [], 0
    for i, r in enumerate(rows[1:], start=2):           # row 1 = header
        account = _to_int_account(_cell(r, 0))
        if account is None:
            continue                                    # blank padding row
        line += 1
        events.append(_normalise(
            account=account,
            date=_to_date(_cell(r, 2)),
            debet=_cell(r, 5), credit=_cell(r, 6),
            doc_no=_doc_no(_cell(r, 3)),
            journal=str(_cell(r, 8)).strip() if _cell(r, 8) else "",
            description=str(_cell(r, 7)).strip() if _cell(r, 7) else "",
            counterparty=_doc_no(_cell(r, 4)) if _cell(r, 4) is not None else None,
            source_system=source_system, source_file=source_file,
            excel_row=i, line_no=line, gross_up=gross_up,
        ))
    return events, None


# ----------------------------------------------------------------------------
# ingestion + reconciliation
# ----------------------------------------------------------------------------

def ingest_file(path, gross_up=APPLY_VAT_GROSS_UP):
    rows = _rows(path)
    fname = os.path.basename(path)
    fmt = detect_format(rows)
    system = SOURCE_SYSTEM_BY_FILE.get(fname, SOURCE_SYSTEM_BY_FORMAT[fmt])
    if fmt == "fintransactions":
        events, eindsaldo = parse_fintransactions(rows, system, fname, gross_up)
    else:
        events, eindsaldo = parse_gb(rows, system, fname, gross_up)

    # reconciliation: our net total must equal the file's own Eindsaldo
    # (only Exact files carry one). This is the "no two numbers disagree" check.
    net_sum = round(sum(e["net_amount"] for e in events), 2)
    recon = {
        "file": fname, "format": fmt, "source_system": system,
        "events": len(events), "net_sum": net_sum, "eindsaldo": eindsaldo,
        "reconciles": (eindsaldo is None) or (abs(net_sum - eindsaldo) < 0.01),
    }
    return events, recon


def ingest_folder(folder, gross_up=APPLY_VAT_GROSS_UP):
    all_events, recon = [], []
    for path in sorted(glob.glob(os.path.join(folder, "*.xlsx"))):
        ev, rc = ingest_file(path, gross_up=gross_up)
        all_events.extend(ev)
        recon.append(rc)

    # guarantee globally-unique ids / source_row_ids
    seen = defaultdict(int)
    for e in all_events:
        seen[e["source_row_id"]] += 1
        if seen[e["source_row_id"]] > 1:
            suffix = f".{seen[e['source_row_id']]}"
            e["source_row_id"] += suffix
            e["event_id"] += suffix
    return all_events, recon


# ----------------------------------------------------------------------------
# windowing: full timeline -> a 13-week horizon array (week 1..13)
# ----------------------------------------------------------------------------

def window_to_horizon(events, anchor=None, horizon=13):
    """Slice the events to a `horizon`-week window starting at `anchor`
    (a date). Stamps week 1..13 and returns the CASH_EVENTS-shaped array.
    Default anchor = earliest transaction date, so output is non-empty out of
    the box; in production set anchor to your forecast start (e.g. today)."""
    dated = [e for e in events if e["date"]]
    if not dated:
        return []
    if anchor is None:
        anchor = min(dt.date.fromisoformat(e["date"]) for e in dated)
    elif isinstance(anchor, str):
        anchor = dt.date.fromisoformat(anchor)

    out = []
    for e in dated:
        d = dt.date.fromisoformat(e["date"])
        wk = (d - anchor).days // 7 + 1
        if 1 <= wk <= horizon:
            ev = dict(e)
            ev["week"] = wk
            out.append(ev)
    return out


# ----------------------------------------------------------------------------
# aggregation + audit helpers — 1:1 port of the JS module. Every dashboard
# number is COMPUTED from the events, never stored. This is the source of truth.
# ----------------------------------------------------------------------------

DRIVERS = ["materials", "subcontractor", "milestone_billing", "payment_lag",
           "weather", "unmapped"]


def eventsForScenario(all_events, scenario):
    if scenario == "base":
        return [e for e in all_events if e["scenario"] == "base"]
    overrides = [e for e in all_events if e["scenario"] == scenario]
    overridden = {e["source_row_id"] for e in overrides}
    base_kept = [e for e in all_events
                 if e["scenario"] == "base" and e["source_row_id"] not in overridden]
    return base_kept + overrides


def weeklyByDriver(events, horizon=13):
    weeks = {w: {d: 0.0 for d in DRIVERS} | {"net": 0.0} for w in range(1, horizon + 1)}
    for e in events:
        w = e.get("week")
        if w in weeks:
            weeks[w][e["driver"]] += e["amount"]
            weeks[w]["net"] += e["amount"]
    return weeks


def runningBalance(events, opening=OPENING_CASH, horizon=13):
    weekly = weeklyByDriver(events, horizon)
    out, bal = [], opening
    for w in range(1, horizon + 1):
        bal += weekly[w]["net"]
        out.append({
            "week": w, "net": round(weekly[w]["net"], 2), "balance": round(bal, 2),
            "breach": bal < COVENANT["threshold"],
            "warning": COVENANT["threshold"] <= bal < COVENANT["threshold"] + COVENANT["warningBuffer"],
        })
    return out


def traceFigure(events, week=None, driver=None, opco=None):
    """THE audit trail. Click any figure -> get the exact rows behind it."""
    return [e for e in events
            if (week is None or e.get("week") == week)
            and (driver is None or e["driver"] == driver)
            and (opco is None or e["opco"] == opco)]


def opcoExposure(events):
    by = {}
    for e in events:
        o = by.setdefault(e["opco"], {"inflows": 0.0, "outflows": 0.0, "net": 0.0})
        if e["amount"] >= 0:
            o["inflows"] += e["amount"]
        else:
            o["outflows"] += e["amount"]
        o["net"] += e["amount"]
    return by


# ----------------------------------------------------------------------------
# CLI
# ----------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="Ingest accounting exports -> unified cash_events")
    ap.add_argument("--in", dest="indir", default="/mnt/user-data/uploads")
    ap.add_argument("--out", dest="outdir", default="/mnt/user-data/outputs")
    ap.add_argument("--anchor", default=None, help="forecast start date YYYY-MM-DD (default: earliest txn)")
    ap.add_argument("--no-gross-up", action="store_true", help="keep net ex-VAT figures")
    args = ap.parse_args()
    os.makedirs(args.outdir, exist_ok=True)

    if os.path.isfile(args.indir):
        ev, rc = ingest_file(args.indir, gross_up=not args.no_gross_up)
        events, recon = ev, [rc]
    else:
        events, recon = ingest_folder(args.indir, gross_up=not args.no_gross_up)
    horizon = window_to_horizon(events, anchor=args.anchor)

    # ---- write outputs ----
    full_path = os.path.join(args.outdir, "events_normalized.json")
    cash_path = os.path.join(args.outdir, "cash_events.json")
    with open(full_path, "w", encoding="utf-8") as f:
        json.dump(events, f, ensure_ascii=False, indent=2)
    with open(cash_path, "w", encoding="utf-8") as f:
        json.dump(horizon, f, ensure_ascii=False, indent=2)

    # ---- console report ----
    print("RECONCILIATION  (net = credit - debit, vs the file's own Eindsaldo)")
    print("-" * 78)
    for r in recon:
        flag = "PASS" if r["reconciles"] else "FAIL"
        eind = f"{r['eindsaldo']:>16,.2f}" if r["eindsaldo"] is not None else "      (no total)"
        print(f"  [{flag}] {r['file'][:46]:46} {r['source_system']:9} "
              f"rows={r['events']:>5}  net={r['net_sum']:>16,.2f}  eindsaldo={eind}")
    print("-" * 78)
    print(f"  total normalised events : {len(events):,}")
    print(f"  events in 13-wk horizon : {len(horizon):,}  (anchor "
          f"{args.anchor or 'earliest txn'})")
    by_sys = defaultdict(int)
    by_drv = defaultdict(int)
    for e in events:
        by_sys[e["source_system"]] += 1
        by_drv[e["driver"]] += 1
    print(f"  by source_system        : {dict(by_sys)}")
    print(f"  by driver               : {dict(by_drv)}")
    print(f"\n  wrote {cash_path}  (drop-in CASH_EVENTS, week 1..13)")
    print(f"  wrote {full_path}  (full lossless timeline w/ audit fields)")


if __name__ == "__main__":
    main()
