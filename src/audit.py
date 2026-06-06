"""AUDITABILITY — trace any dashboard figure to its literal source Excel cell.

drill_down()      -> the forecast_events behind a (week, vat_category) figure
trace_seed_rows() -> the historical revenue_actuals rows that seeded a forecast event
read_excel_row()  -> the literal cell values of one source .xlsx row
"""
from __future__ import annotations

import glob
import os

from openpyxl import load_workbook

from . import config

_FINTRANS_COLS = ["Nr.", "Per.", "Datum", "Bkst.nr.", "Dagboek", "Debet", "Credit"]


def drill_down(forecast_events, week=None, vat_category=None):
    """Filter forecast_events to the rows behind a dashboard cell."""
    df = forecast_events
    if week is not None:
        df = df[df["week"] == week]
    if vat_category is not None:
        df = df[df["vat_category"] == vat_category]
    return df.copy()


def trace_seed_rows(event, actuals):
    """Return the revenue_actuals rows that seeded one forecast event.

    Returns an empty frame if the event carries no seed_event_ids.
    """
    ids = event.get("seed_event_ids", [])
    return actuals[actuals["event_id"].isin(ids)].copy()


def read_excel_row(source_file, source_excel_row, glob_pattern=None):
    """Open the originating .xlsx and return that row's labelled cell values.

    `row` is None when the file cannot be found; callers must handle that.
    An out-of-range row index yields a row dict whose cells are all None.
    """
    glob_pattern = config.ACTUAL_DATA_GLOB if glob_pattern is None else glob_pattern
    path = next((p for p in glob.glob(glob_pattern, recursive=True)
                 if os.path.basename(p) == source_file), None)
    if path is None:
        return {"raw_file": source_file, "key": source_excel_row, "row": None}
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    idx = source_excel_row - 1          # 1-based -> 0-based
    cells = rows[idx] if 0 <= idx < len(rows) else ()
    row = {}
    for i, col in enumerate(_FINTRANS_COLS):
        row[col] = cells[i] if i < len(cells) else None
    return {"raw_file": source_file, "key": source_excel_row, "row": row}
