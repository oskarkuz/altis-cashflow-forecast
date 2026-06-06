"""Excel ingestion — Exact FinTransactions exports -> revenue_actuals DataFrame.

Ported from notebooks/ingest.py (the parser is already reconciliation-verified
against the real Dakdekkersbedrijf files). Single company, revenue accounts only.
Every row keeps source_file + source_excel_row so any figure traces to its cell.
"""
from __future__ import annotations

import datetime as dt
import glob
import os
import warnings

import pandas as pd
from openpyxl import load_workbook

from . import config


def _rows(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        return list(ws.iter_rows(values_only=True))
    finally:
        wb.close()


def _cell(row, i):
    return row[i] if row and len(row) > i else None


def _num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0


def _to_int_account(v):
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return None


def _to_date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    return None


def _doc_no(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def classify(account):
    """account -> (vat_category, vat_rate, label). Unknown does not crash."""
    if account in config.GL_ACCOUNTS:
        return config.GL_ACCOUNTS[account]
    return ("unmapped", 0.0, f"ONGEMAPT grootboek {account} — map in config.GL_ACCOUNTS")


def parse_fintransactions(rows, source_file):
    """Return (list[dict] rows, eindsaldo). One dict per ledger line."""
    account = None
    for r in rows[:12]:
        if str(_cell(r, 0)).strip().lower().startswith("grootboekrekening"):
            account = _to_int_account(str(_cell(r, 1)).split(" - ")[0])
            break
    if account is None:
        warnings.warn(f"{source_file}: no 'Grootboekrekening' account header found "
                      "(file format mismatch?)")
    h = next((i for i, r in enumerate(rows) if str(_cell(r, 0)).strip() == "Nr."), None)
    if h is None:
        return [], None
    out, line, eindsaldo = [], 0, None
    for i, r in enumerate(rows[h + 1:], start=h + 2):   # +2 -> 1-based excel row
        first = _cell(r, 0)
        s = "" if first is None else str(first).strip()
        if s in ("", "None", "Totaal"):
            continue
        if s == "Eindsaldo":
            eindsaldo = round(_num(_cell(r, 6)) - _num(_cell(r, 5)), 2)
            continue
        line += 1
        vat_category, vat_rate, label = classify(account)
        net = round(_num(_cell(r, 6)) - _num(_cell(r, 5)), 2)
        cash = round(net * (1 + vat_rate), 2)
        date = _to_date(_cell(r, 2))
        out.append({
            "event_id": f"EX-{_doc_no(_cell(r, 3))}#L{line}",
            "date": date,
            "gl_account": account,
            "vat_category": vat_category,
            "vat_rate": vat_rate,
            "label": label,
            "net_amount": net,
            "cash_amount": cash,
            "debet": round(_num(_cell(r, 5)), 2),
            "credit": round(_num(_cell(r, 6)), 2),
            "doc_no": _doc_no(_cell(r, 3)),
            "journal": str(_cell(r, 4)).strip() if _cell(r, 4) else "",
            "source_file": source_file,
            "source_excel_row": i,
        })
    return out, eindsaldo


def load_revenue_actuals(glob_pattern=None):
    """Load + reconcile all matching files. Returns (actuals_df, recon_report)."""
    glob_pattern = glob_pattern or config.ACTUAL_DATA_GLOB
    all_rows, recon = [], []
    for path in sorted(glob.glob(glob_pattern, recursive=True)):
        fname = os.path.basename(path)
        rows = _rows(path)
        parsed, eindsaldo = parse_fintransactions(rows, fname)
        all_rows.extend(parsed)
        net_sum = round(sum(r["net_amount"] for r in parsed), 2)
        recon.append({
            "file": fname, "rows": len(parsed), "net_sum": net_sum,
            "eindsaldo": eindsaldo,
            "reconciles": eindsaldo is None or abs(net_sum - eindsaldo) < 0.01,
        })
        if not recon[-1]["reconciles"]:
            warnings.warn(f"{fname}: net_sum {net_sum} != eindsaldo {eindsaldo} "
                          "(does not reconcile)")

    df = pd.DataFrame(all_rows)
    if df.empty:
        df = pd.DataFrame(columns=[
            "event_id", "date", "gl_account", "vat_category", "vat_rate", "label",
            "net_amount", "cash_amount", "debet", "credit", "doc_no", "journal",
            "source_file", "source_excel_row", "iso_year", "iso_week"])
        return df, recon

    # globally-unique event_id (doc/line collisions across files)
    dup = df["event_id"].duplicated(keep=False)
    df.loc[dup, "event_id"] = (df.loc[dup, "event_id"] + "@"
                               + df.loc[dup, "source_file"]
                               + "#" + df.loc[dup, "source_excel_row"].astype(str))
    iso = df["date"].map(lambda d: d.isocalendar() if d else (0, 0, 0))
    df["iso_year"] = iso.map(lambda c: c[0])
    df["iso_week"] = iso.map(lambda c: c[1])
    return df, recon
